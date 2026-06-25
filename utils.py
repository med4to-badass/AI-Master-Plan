"""
utils.py
Funções utilitárias:
- Geração de mensagens prontas para WhatsApp (profissionais e amigáveis)
- Exportação de relatórios para Excel com formatação bonita
- Formatação de datas e valores para o Brasil
"""

from datetime import date, datetime, timezone, timedelta

_BRT = timezone(timedelta(hours=-3))
from io import BytesIO
from typing import Any, Dict, List

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

from calculations import (
    format_currency,
    get_milestone_progress,
)


# ================== WHATSAPP MESSAGES ==================

def create_welcome_purchase_message(
    client_name: str,
    amount: float,
    final_points: int,
    current_points: int,
    available_packages: int,
    client_link: str = "",
) -> str:
    """Mensagem simples após compra (1 pacote = 1 ponto). Sem resgate."""
    first = client_name.split()[0] if client_name else "Parceiro"

    msg = f"""Olá {first}!

+{final_points} ponto(s) por esta compra.

Saldo atual: {current_points} pontos

Continue comprando para chegar aos 500 pontos e ganhar a Cafeteira!

Obrigado!"""

    return msg


def create_redemption_message(
    client_name: str,
    packages: int,
    points_deducted: int,
    remaining_points: int
) -> str:
    """Mensagem de confirmação de resgate de pacote."""
    pkg_word = "pacote" if packages == 1 else "pacotes"

    msg = f"""Parabéns {client_name.split()[0]}! 🎉

Seu resgate foi realizado com sucesso:

📦 *{packages} {pkg_word} grátis* resgatado(s)
   ({points_deducted} pontos utilizados)

Seu novo saldo: *{remaining_points} pontos*

O pacote está separado e pode ser retirado quando quiser. 
Qualquer espessura! É só avisar.

Obrigado por ser um parceiro fiel do IsoSoluções! 💚

_IsoSoluções • Programa Parceiro Isopor_"""

    return msg


def create_monthly_summary_message(
    client_name: str,
    monthly_spent: float,
    points_this_month: int,
    available_packages: int = 0,
) -> str:
    """Resumo simples do mês (sem resgate)."""
    spent_str = format_currency(monthly_spent)

    msg = f"""Resumo do seu mês no Programa Parceiro Isopor, {client_name.split()[0]}!

📅 Pacotes comprados: *{spent_str}*
⭐ Pontos ganhos: *{points_this_month}*

Continue comprando para chegar aos 500 pontos e ganhar a Cafeteira!

Obrigado!
Equipe IsoSoluções 💚"""

    return msg


def create_promotional_message() -> str:
    """Mensagem geral de divulgação do programa (para usar em grupos ou novos clientes) — sem bônus mensal."""
    return """🌟 *Programa Parceiro Isopor* - IsoSoluções 🌟

A cada R$ 38 em compras = 1 ponto
A cada 10 pontos = 1 pacote grátis (qualquer espessura!)

Seus pontos são cumulativos: cada ponto conquistado se soma ao saldo anterior.

Quanto mais você compra, mais pacotes grátis resgata!

Quer participar? É só falar comigo que eu já cadastro você.

_IsoSoluções • Qualidade e fidelidade que você merece_ 💚"""


def create_milestone_reward_message(
    client_name: str,
    reward_description: str = "Cafeteira",
    program_name: str = "Programa Parceiro Isopor",
) -> str:
    """Mensagem simples quando bate 500 pontos."""
    first = client_name.split()[0] if client_name else "Parceiro"
    msg = f"""Parabéns {first}! 🏆

Você chegou a 500 pontos!

Brinde: **{reward_description}**

Pode retirar quando quiser. Obrigado pela confiança! 💚"""
    return msg


# ================== EXCEL EXPORT ==================

def _style_header(ws, row_num: int, fill_color: str = "0D9488"):
    """Aplica estilo profissional ao cabeçalho."""
    fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
    font = Font(bold=True, color="FFFFFF", size=11)
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='334155'),
        right=Side(style='thin', color='334155'),
        top=Side(style='thin', color='334155'),
        bottom=Side(style='thin', color='334155')
    )

    for cell in ws[row_num]:
        cell.fill = fill
        cell.font = font
        cell.alignment = alignment
        cell.border = thin_border


def _auto_column_width(ws, min_width: int = 12, max_width: int = 45):
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted


def export_full_report(
    clients: List[Dict[str, Any]],
    kpis: Dict[str, Any],
    monthly_history: List[Dict[str, Any]],
    top_clients: List[Dict[str, Any]],
    recent_activity: List[Dict[str, Any]]
) -> BytesIO:
    """
    Gera um arquivo Excel completo e bem formatado com várias abas.
    Retorna BytesIO pronto para download no Streamlit.
    """
    wb = Workbook()

    # ================== ABA 1: RESUMO EXECUTIVO ==================
    ws = wb.active
    ws.title = "Resumo Executivo"

    # Título
    ws.merge_cells("A1:F1")
    ws["A1"] = "📊 RELATÓRIO COMPLETO - PROGRAMA PARCEIRO ISOPOR"
    ws["A1"].font = Font(bold=True, size=16, color="10b981")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:F2")
    ws["A2"] = f"IsoSoluções • Gerado em {datetime.now(tz=_BRT).strftime('%d/%m/%Y às %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="64748b")
    ws["A2"].alignment = Alignment(horizontal="center")

    # KPIs
    ws["A4"] = "INDICADORES PRINCIPAIS"
    ws["A4"].font = Font(bold=True, size=12, color="0f172a")

    kpi_data = [
        ("Total de Clientes", kpis["total_clients"]),
        ("Pontos Distribuídos Hoje", kpis["points_today"]),
        ("Pacotes Resgatados Este Mês", kpis["packages_redeemed_month"]),
        ("Volume de Compras Este Mês (R$)", round(kpis["volume_month"], 2)),
        ("Pontos em Circulação (Saldo)", kpis["circulating_points"]),
        ("Pontos em Circulação (Total)", kpis["circulating_points"]),
    ]

    for i, (label, value) in enumerate(kpi_data, start=5):
        ws[f"A{i}"] = label
        ws[f"B{i}"] = value
        ws[f"A{i}"].font = Font(bold=True)
        ws[f"B{i}"].font = Font(bold=True, color="10b981")

    # ================== ABA 2: CLIENTES ATUAIS ==================
    ws2 = wb.create_sheet("Clientes Atuais")

    df_clients = pd.DataFrame([
        {
            "Nome": c["name"],
            "Telefone": c["phone"],
            "Pontos Atuais": c["current_points"],
            "Pacotes Disponíveis": c["available_packages"],
            "Pacotes Comprados (Total)": c.get("total_packages_bought", 0),
            "Volume Este Mês (R$)": round(c["monthly_spent"], 2),
            "Total Gasto (R$)": round(c["total_spent"], 2),
            "Última Compra": c.get("last_purchase_date", "—") or "—",
            "Recompensa 500?": "Sim" if c.get("has_milestone_500") else "—",
        }
        for c in clients
    ])

    for r_idx, row in enumerate(dataframe_to_rows(df_clients, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws2.cell(row=r_idx, column=c_idx, value=value)

    _style_header(ws2, 1, "0f766e")
    _auto_column_width(ws2)

    # ================== ABA 3: HISTÓRICO MENSAL ==================
    ws3 = wb.create_sheet("Histórico Mensal")

    df_monthly = pd.DataFrame(monthly_history)
    if not df_monthly.empty:
        df_monthly = df_monthly.rename(columns={
            "month": "Mês (Ano-Mês)",
            "volume": "Volume Compras (R$)",
            "points_awarded": "Pontos Concedidos",
            "num_purchases": "Qtd. Compras"
        })

    for r_idx, row in enumerate(dataframe_to_rows(df_monthly, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws3.cell(row=r_idx, column=c_idx, value=value)

    _style_header(ws3, 1, "1e40af")
    _auto_column_width(ws3)

    # ================== ABA 4: TOP 10 FIÉIS ==================
    ws4 = wb.create_sheet("Top 10 Mais Fiéis")

    df_top = pd.DataFrame([
        {
            "Posição": idx + 1,
            "Nome": c["name"],
            "Telefone": c["phone"],
            "Pontos Totais Ganhos": c["total_points_earned"],
            "Saldo Atual de Pontos": c["current_points"],
            "Pacotes Disponíveis": c["available_packages"],
            "Pacotes Comprados": c.get("total_packages_bought", 0),
        }
        for idx, c in enumerate(top_clients)
    ])

    for r_idx, row in enumerate(dataframe_to_rows(df_top, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            ws4.cell(row=r_idx, column=c_idx, value=value)

    _style_header(ws4, 1, "854d0e")
    _auto_column_width(ws4)

    # ================== ABA 5: ATIVIDADE RECENTE ==================
    ws5 = wb.create_sheet("Atividade Recente")

    activity_rows = []
    for act in recent_activity:
        if act["type"] == "purchase":
            activity_rows.append({
                "Data": act["tx_date"],
                "Tipo": "COMPRA",
                "Cliente": act["client_name"],
                "Valor (R$)": round(act["amount"], 2) if act["amount"] else 0,
                "Pontos": act["points"],
                "Observação": act["notes"] or ""
            })
        elif act["type"] == "redemption":
            activity_rows.append({
                "Data": act["tx_date"],
                "Tipo": "RESGATE",
                "Cliente": act["client_name"],
                "Valor (R$)": 0,
                "Pontos": act["points"],
                "Observação": f"Resgate de pacote(s) - {act['notes'] or ''}"
            })
        else:
            # milestone_reward
            activity_rows.append({
                "Data": act["tx_date"],
                "Tipo": "RECOMPENSA ESPECIAL",
                "Cliente": act["client_name"],
                "Valor (R$)": 0,
                "Pontos": 0,
                "Observação": act["notes"] or "Recompensa marco 500 pacotes"
            })

    df_act = pd.DataFrame(activity_rows)
    for r_idx, row in enumerate(dataframe_to_rows(df_act, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws5.cell(row=r_idx, column=c_idx, value=value)
            if r_idx > 1:
                if value == "COMPRA":
                    cell.fill = PatternFill(start_color="dcfce7", end_color="dcfce7", fill_type="solid")
                elif value == "RESGATE":
                    cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")

    _style_header(ws5, 1, "334155")
    _auto_column_width(ws5)

    # Salvar em memória
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
